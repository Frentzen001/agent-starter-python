"""
knowledge.py — Garage@EEE knowledge base for More-Tea.

Loads structured information from the xlsx database at import time and
exposes a single ``GARAGE_KNOWLEDGE`` string for injection into the
agent's system prompt.  The xlsx file path can be overridden via the
GARAGE_XLSX environment variable; if the file is not found the module
falls back to a minimal hardcoded summary so the agent can still run.
"""

from __future__ import annotations

import logging
import os
from pathlib import Path

logger = logging.getLogger("knowledge")

# ---------------------------------------------------------------------------
# Locate the xlsx
# ---------------------------------------------------------------------------
_DEFAULT_XLSX = Path(__file__).parent / "Garage@EEEWebsiteDatabase.xlsx"
_XLSX_PATH = Path(os.environ.get("GARAGE_XLSX", str(_DEFAULT_XLSX)))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _clean(value: object) -> str:
    """Return a clean single-line string, or '' if None/blank."""
    if value is None:
        return ""
    s = str(value).strip()
    # Collapse excessive newlines / indentation from xlsx cells
    lines = [ln.strip() for ln in s.splitlines() if ln.strip()]
    return " ".join(lines)


def _rows(ws) -> list[tuple]:
    """Return all rows as tuples, skipping wholly-empty rows."""
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v for v in row if v is not None and str(v).strip()):
            result.append(row)
    return result


# ---------------------------------------------------------------------------
# Sheet extractors
# ---------------------------------------------------------------------------
def _extract_home(wb) -> str:
    ws = wb["Home"]
    cols = {cell.value: i for i, cell in enumerate(next(ws.iter_rows(max_row=1)))}
    data = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
    if data is None:
        return ""
    lines = ["## About Garage@EEE"]
    if (i := cols.get("About")) is not None and data[i]:
        lines.append(_clean(data[i]))
    if (i := cols.get("Objective")) is not None and data[i]:
        lines.append(f"\nObjective: {_clean(data[i])}")
    if (i := cols.get("Facilities")) is not None and data[i]:
        lines.append(f"\nFacilities overview: {_clean(data[i])}")
    return "\n".join(lines)


def _extract_facilities(wb) -> str:
    ws = wb["Facilities"]
    lines = ["## Facilities"]
    for row in _rows(ws):
        name, desc = _clean(row[0]), _clean(row[1])
        if name and desc and name.lower() != "cover pic":
            lines.append(f"- {name}: {desc}")
    return "\n".join(lines)


def _extract_events(wb) -> str:
    ws = wb["Events"]
    lines = ["## Events & Programmes"]
    for row in _rows(ws):
        name, tagline, desc = _clean(row[0]), _clean(row[1]), _clean(row[2])
        if name:
            entry = f"- {name}"
            if tagline:
                entry += f" ({tagline})"
            if desc:
                entry += f": {desc[:300]}{'...' if len(desc) > 300 else ''}"
            lines.append(entry)
    return "\n".join(lines)


def _extract_ambassadors(wb) -> str:
    ws = wb["Ambassadors"]
    lines = ["## Ambassador Portfolios"]
    for row in _rows(ws):
        name, desc = _clean(row[0]), _clean(row[1])
        if name and desc:
            lines.append(f"- {name}: {desc}")
    return "\n".join(lines)


def _extract_programs(wb) -> str:
    """Tinkering Projects, Innotrack, and Launchpad."""
    sections = []

    # Tinkering Projects
    ws = wb["Tinkering Project"]
    rows = _rows(ws)
    if rows:
        intro = _clean(rows[0][2]) if rows[0][2] else ""
        how = _clean(rows[0][3]) if rows[0][3] else ""
        funding = ""
        faqs = []
        for row in rows:
            if _clean(row[9]) == "Funding" and row[10]:
                funding = _clean(row[10])
            q, a = _clean(row[16]), _clean(row[17])
            if q and a:
                faqs.append(f"  Q: {q}\n  A: {a}")
        block = "## Tinkering Projects\n"
        if intro:
            block += f"{intro}\n"
        if how:
            block += f"How to join: {how}\n"
        if funding:
            block += f"Funding: {funding}\n"
        if faqs:
            block += "FAQs:\n" + "\n".join(faqs)
        sections.append(block)

    # Innotrack
    ws = wb["Innotrack"]
    rows = _rows(ws)
    if rows:
        funding = ""
        faqs = []
        for row in rows:
            if _clean(row[9]) == "Funding" and row[10]:
                funding = _clean(row[10])
            q, a = _clean(row[14]), _clean(row[15])
            if q and a:
                faqs.append(f"  Q: {q}\n  A: {a}")
        block = "## Innovators Track (Innotrack)\n"
        block += (
            "A programme for students with start-up ideas. Teams can propose projects, "
            "receive mentorship, industry connections, and access to facilities.\n"
        )
        if funding:
            block += f"Funding: {funding}\n"
        if faqs:
            block += "FAQs:\n" + "\n".join(faqs)
        sections.append(block)

    # Launchpad
    ws = wb["Launchpad"]
    rows = _rows(ws)
    if rows:
        funding = ""
        faqs = []
        for row in rows:
            if len(row) > 6 and _clean(row[6]) in ("Funding", "Funding ") and row[7]:
                funding = _clean(row[7])
            if len(row) > 9:
                q, a = _clean(row[8]), _clean(row[9])
                if q and a:
                    faqs.append(f"  Q: {q}\n  A: {a}")
        block = "## Launchpad\n"
        block += (
            "Launchpad supports student start-ups with funding, mentorship, industry "
            "connections, and makerspace access. Open to all NTU students with at least "
            "one EEE student per team.\n"
        )
        if funding:
            block += f"Funding: {funding}\n"
        if faqs:
            block += "FAQs:\n" + "\n".join(faqs)
        sections.append(block)

    return "\n\n".join(sections)


def _extract_current_projects(wb) -> str:
    ws = wb["Project Openings"]
    lines = ["## Current Projects at Garage@EEE"]
    for row in _rows(ws):
        name, desc = _clean(row[0]), _clean(row[2])
        programme = _clean(row[3])
        recruiting = _clean(row[6])
        if name and desc:
            entry = f"- {name}"
            if programme:
                entry += f" [{programme}]"
            if recruiting and recruiting.upper() == "Y":
                entry += " [RECRUITING]"
            entry += f": {desc[:200]}{'...' if len(desc) > 200 else ''}"
            lines.append(entry)
    return "\n".join(lines)


def _extract_past_projects(wb) -> str:
    ws = wb["Project Info"]
    lines = ["## Notable Past Projects"]
    for row in _rows(ws):
        name, tagline, desc = _clean(row[0]), _clean(row[4]), _clean(row[5])
        if name and (tagline or desc):
            entry = f"- {name}"
            if tagline:
                entry += f" — {tagline}"
            if desc:
                entry += f": {desc[:200]}{'...' if len(desc) > 200 else ''}"
            lines.append(entry)
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Build the knowledge string
# ---------------------------------------------------------------------------
def _build_knowledge() -> str:
    if not _XLSX_PATH.exists():
        logger.warning(
            f"Garage xlsx not found at {_XLSX_PATH} — using minimal fallback knowledge."
        )
        return (
            "Garage@EEE is a student-led makerspace in NTU's School of Electrical and "
            "Electronic Engineering. It provides equipment, funding, and mentorship for "
            "student projects. Key programmes: Tinkering Projects (up to $200 funding), "
            "Innotrack (up to $999), Launchpad (up to $999). Facilities include 3D "
            "printers, laser cutter, PCB milling machine, soldering station, electronics "
            "workstation, drill press, acrylic bender, and workbench."
        )

    try:
        import openpyxl  # soft dependency — only needed at startup

        wb = openpyxl.load_workbook(str(_XLSX_PATH), data_only=True)
        sections = [
            _extract_home(wb),
            _extract_facilities(wb),
            _extract_events(wb),
            _extract_ambassadors(wb),
            _extract_programs(wb),
            _extract_current_projects(wb),
            _extract_past_projects(wb),
        ]
        knowledge = "\n\n".join(s for s in sections if s.strip())
        logger.info(
            f"Garage knowledge loaded from xlsx ({len(knowledge)} chars)."
        )
        return knowledge
    except Exception as exc:
        logger.error(f"Failed to load Garage knowledge from xlsx: {exc}")
        return "Garage@EEE is a student-led makerspace in NTU EEE providing facilities, funding, and mentorship for student innovation projects."


# Eagerly built at import time — zero latency at query time
GARAGE_KNOWLEDGE: str = _build_knowledge()
